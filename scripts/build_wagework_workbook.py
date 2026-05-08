import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT = Path("outputs/wagework")
RAW_PATH = OUT / "wagework_selected_wage_results_raw.json"
ALL_RAW_PATH = OUT / "wagework_all_occupation_wage_results_raw.json"
XLSX_PATH = OUT / "워크피디아_맞춤형임금정보_직무별_정리본.xlsx"


def as_num(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", ""))
    except ValueError:
        return None


def extract_rows(path=RAW_PATH):
    raw = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for rec in raw:
        result_list = ((rec.get("result") or {}).get("data") or {}).get("resultList")
        if not isinstance(result_list, dict):
            rows.append(
                {
                    "앱분류": rec.get("group"),
                    "직업코드": rec.get("occpClCd"),
                    "직업명": rec.get("occpCfnm"),
                    "분류수준": rec.get("strfLvlSn"),
                    "상위코드": rec.get("upprCdId"),
                    "결과상태": "결과 없음",
                }
            )
            continue

        avg = as_num(result_list.get("entrAvrgAnslAmt"))
        median = as_num(result_list.get("midAvrgIcomAmt"))
        p25 = as_num(result_list.get("lwprAvrgIcomAmt"))
        p75 = as_num(result_list.get("upprAvrgIcomAmt"))
        sample_workers = as_num(result_list.get("ivchLbrn"))
        population_workers = as_num(result_list.get("prspLbrn"))

        rows.append(
            {
                "앱분류": rec.get("group"),
                "직업코드": rec.get("occpClCd"),
                "직업명": rec.get("occpCfnm"),
                "분류수준": rec.get("strfLvlSn"),
                "상위코드": rec.get("upprCdId"),
                "직업분류경로": result_list.get("occpFullCdNm")
                or result_list.get("occpLrclCdNm")
                or result_list.get("occpMlsfCdNm")
                or result_list.get("occpSclaCdNm")
                or rec.get("occpCfnm"),
                "평균연간임금_천원": avg,
                "월평균환산_만원": round(avg / 12 / 10, 1) if avg else None,
                "하위25_연간임금_천원": p25,
                "중위_연간임금_천원": median,
                "상위25_연간임금_천원": p75,
                "상대표준오차": as_num(result_list.get("rltlSd")),
                "표본근로자수": sample_workers,
                "추정근로자수": population_workers,
                "결과상태": "성공",
            }
        )
    return rows


def choose_app_mapping(rows):
    pick_codes = {
        "사무": ["312", "313", "314", "399"],
        "마케팅": ["273"],
        "데이터": ["223"],
        "개발": ["222"],
        "디자인": ["285"],
    }
    result = []
    by_code = {str(r.get("직업코드")): r for r in rows if r.get("결과상태") == "성공"}
    for group, codes in pick_codes.items():
        chosen = [by_code[c] for c in codes if c in by_code]
        if not chosen:
            continue
        avg = sum(r["평균연간임금_천원"] for r in chosen) / len(chosen)
        median = sum(r["중위_연간임금_천원"] for r in chosen) / len(chosen)
        result.append(
            {
                "앱직무": group,
                "사용직업코드": ", ".join(codes),
                "사용직업명": " / ".join(r["직업명"] for r in chosen),
                "평균연간임금_천원": round(avg, 0),
                "월평균환산_만원": round(avg / 12 / 10, 1),
                "중위연간임금_천원": round(median, 0),
                "중위월환산_만원": round(median / 12 / 10, 1),
                "앱반영메모": "내일경로 AI 임금 기대값의 직무별 기준선 후보",
            }
        )
    return result


def write_table(ws, rows, start_row=1, title=None):
    if title:
        ws.cell(start_row, 1, title)
        ws.cell(start_row, 1).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(start_row, 1).fill = PatternFill("solid", fgColor="003675")
        start_row += 2
    if not rows:
        return start_row
    headers = list(rows[0].keys())
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start_row, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D56A5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, h in enumerate(headers, 1):
            ws.cell(r_idx, c_idx, row.get(h))
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws.cell(start_row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    for idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), *(len(str(row.get(h) or "")) for row in rows[:200]))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)
    return start_row + len(rows) + 2


def main():
    rows = extract_rows(RAW_PATH)
    all_rows = extract_rows(ALL_RAW_PATH) if ALL_RAW_PATH.exists() else []
    mapping = choose_app_mapping(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "앱용 요약"
    write_table(ws, mapping, title="워크피디아 직무별 임금 기준선")

    ws2 = wb.create_sheet("워크피디아 원자료")
    write_table(ws2, rows, title="워크피디아 맞춤형 임금정보 조회 결과")

    ws_all = wb.create_sheet("전체 직업분류")
    write_table(ws_all, all_rows, title="워크피디아 전체 직업분류 임금정보")

    ws3 = wb.create_sheet("검색조건 및 출처")
    source_rows = [
        {
            "항목": "출처",
            "내용": "임금직업포털 워크피디아 맞춤형 임금정보",
        },
        {
            "항목": "URL",
            "내용": "https://www.wagework.go.kr/pt/c/a/retrieveCstmWageSrch.do",
        },
        {
            "항목": "검색조건",
            "내용": "직업별 조건 1개 선택, 한국표준직업분류 제7차 기준",
        },
        {
            "항목": "수집범위",
            "내용": "사무, 마케팅, 데이터, 개발, 디자인 관련 직업분류 코드 21건",
        },
        {
            "항목": "전체수집",
            "내용": f"워크피디아 직업분류 전체 {len(all_rows)}건도 별도 탭에 수록",
        },
        {
            "항목": "제공연도",
            "내용": "워크피디아 안내 기준 2024년 기준 추정치",
        },
        {
            "항목": "임금정의",
            "내용": "연간 임금, 정액급여와 특별급여 합산, 초과급여 제외",
        },
        {
            "항목": "주의",
            "내용": "직업별 임금 기준선으로 사용하며, 2026년 실제 채용공고 임금은 별도 보정 필요",
        },
    ]
    write_table(ws3, source_rows, title="수집 방식")

    ws4 = wb.create_sheet("검토 필요")
    review_rows = [
        {"구분": "사무", "판단": "세부 직무가 넓어 경영 사무, 회계 경리, 사무 보조, 고객 상담을 따로 둘 수 있음"},
        {"구분": "마케팅", "판단": "관리자 코드는 초년생 경로에 과대평가될 수 있어 상품 기획, 홍보 및 조사 전문가를 우선 권장"},
        {"구분": "데이터", "판단": "워크피디아는 데이터 및 네트워크 관련 전문가 단위라 데이터 분석가만 분리되지는 않음"},
        {"구분": "개발", "판단": "컴퓨터 시스템 및 소프트웨어 전문가를 개발 직무 기준선으로 사용 가능"},
        {"구분": "디자인", "판단": "디자이너 대분류 기준이라 세부 디자인 직무와 함께 보완하면 좋음"},
    ]
    write_table(ws4, review_rows, title="앱 반영 전 판단")

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = Font(name="Malgun Gothic", size=10, bold=cell.font.bold, color=cell.font.color)
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(XLSX_PATH)


if __name__ == "__main__":
    main()
