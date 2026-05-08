from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOURCE_DIR = Path(r"C:\Users\USER\Desktop\데이터 셋 모음 고용노동")
OUT_DIR = Path(__file__).resolve().parents[1] / "outputs" / "user-data-audit"
OUT_JSON = OUT_DIR / "user_moel_dataset_audit.json"


def compact(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def row_values(row):
    return [compact(cell.value) for cell in row]


def nonempty(values):
    return [v for v in values if v]


def find_header_and_samples(ws, max_scan=30):
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, max_scan)):
        values = row_values(row)
        rows.append(values)

    header_idx = None
    best_score = -1
    for idx, values in enumerate(rows, start=1):
        items = nonempty(values)
        score = len(items)
        if any(any(token in cell for token in ["구분", "지역", "직종", "산업", "연월", "년월", "월", "계", "합계"]) for cell in items):
            score += 3
        if score > best_score and len(items) >= 2:
            best_score = score
            header_idx = idx

    headers = rows[header_idx - 1] if header_idx else []
    headers = [h for h in headers if h]
    samples = []
    if header_idx:
        for row in ws.iter_rows(min_row=header_idx + 1, max_row=min(ws.max_row or header_idx, header_idx + 5)):
            values = [v for v in row_values(row) if v]
            if values:
                samples.append(values[:12])
    return header_idx, headers[:40], samples


def extract_periods(ws, max_rows=5000):
    periods = []
    pattern = re.compile(r"(20[0-2][0-9])[\.\-/년\s]*(0?[1-9]|1[0-2])?")
    rows_seen = 0
    for row in ws.iter_rows():
        rows_seen += 1
        if rows_seen > max_rows:
            break
        for cell in row:
            text = compact(cell.value)
            if not text:
                continue
            for match in pattern.finditer(text):
                year = int(match.group(1))
                month = int(match.group(2) or 1)
                periods.append(f"{year:04d}-{month:02d}")
    return sorted(set(periods))


def classify_file(file_name, sheet_summaries):
    name = file_name
    all_headers = " ".join(" ".join(s["headers"]) for s in sheet_summaries)
    text = f"{name} {all_headers}"
    categories = []
    if "구인구직" in text or "유효구인" in text or "취업현황" in text:
        categories.append("labor_demand")
    if "직종" in text:
        categories.append("occupation")
    if "산업" in text:
        categories.append("industry")
    if "지역" in text or "시도" in text:
        categories.append("region")
    if "내일배움" in text or "훈련" in text:
        categories.append("training")
    if "국민취업지원" in text:
        categories.append("counseling_policy")
    if "고용장려금" in text or "장려금" in text:
        categories.append("subsidy")
    if "실업급여" in text:
        categories.append("unemployment")
    if "임금" in text or "급여" in text:
        categories.append("wage_or_benefit")
    if "사업장" in text or "피보험자" in text:
        categories.append("employment_base")
    return sorted(set(categories)) or ["unknown"]


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(SOURCE_DIR.glob("*.xlsx"))
    audit = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceDir": str(SOURCE_DIR),
        "fileCount": len(files),
        "files": [],
    }

    for path in files:
        print(f"감사 중: {path.name}")
        file_info = {
            "fileName": path.name,
            "path": str(path),
            "bytes": path.stat().st_size,
            "sheets": [],
            "status": "ok",
        }
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_row, headers, samples = find_header_and_samples(ws)
                periods = extract_periods(ws)
                file_info["sheets"].append({
                    "sheetName": sheet_name,
                    "maxRow": ws.max_row,
                    "maxColumn": ws.max_column,
                    "headerRow": header_row,
                    "headers": headers,
                    "sampleRows": samples,
                    "periodMin": periods[0] if periods else "",
                    "periodMax": periods[-1] if periods else "",
                    "periodCount": len(periods),
                })
            wb.close()
            file_info["categories"] = classify_file(path.name, file_info["sheets"])
        except Exception as exc:
            file_info["status"] = "error"
            file_info["error"] = str(exc)
            file_info["categories"] = ["error"]
        audit["files"].append(file_info)

    OUT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"saved: {OUT_JSON}")


if __name__ == "__main__":
    main()
