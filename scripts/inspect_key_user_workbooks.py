from pathlib import Path
from openpyxl import load_workbook

from audit_user_moel_datasets import SOURCE_DIR

KEYWORDS = [
    "구인구직취업현황(월)",
    "내일배움카드발급현황",
    "국민취업지원제도",
    "직종별_구인구직",
    "산업_규모별_구인구직",
    "훈련기관과정현황",
    "실업자훈련취업률",
]


def main():
    for key in KEYWORDS:
        matches = [p for p in SOURCE_DIR.glob("*.xlsx") if key in p.name]
        for path in matches[:1]:
            print("\n====", path.name, "====")
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            print("sheet", ws.title, "rows", ws.max_row, "cols", ws.max_column)
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), start=1):
                values = []
                for value in row[:18]:
                    text = "" if value is None else str(value).replace("\n", " ")
                    values.append(text[:45])
                print(idx, values)
            wb.close()


if __name__ == "__main__":
    main()
