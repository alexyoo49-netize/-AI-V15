from __future__ import annotations

import csv
import json
import math
import re
import zipfile
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE = Path(__file__).resolve().parents[1]
RAW_DIR = BASE / "public" / "data" / "raw"
DATASETS_JSON = BASE / "public" / "data" / "datasets.json"
SUMMARY_JSON = BASE / "public" / "data" / "moel_analysis_summary.json"

REGIONS = ["서울", "경기", "부산", "대구", "광주", "대전", "인천", "울산", "세종", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주"]
REGION_ALIASES = {
    "서울": ["서울"],
    "경기": ["경기", "수원", "성남", "안양", "부천", "고양", "용인", "의정부"],
    "부산": ["부산"],
    "대구": ["대구"],
    "광주": ["광주"],
    "대전": ["대전"],
    "인천": ["인천"],
    "울산": ["울산"],
    "세종": ["세종"],
    "강원": ["강원", "춘천", "원주", "강릉"],
    "충북": ["충북", "청주", "충청북도"],
    "충남": ["충남", "천안", "충청남도"],
    "전북": ["전북", "전라북도", "전북특별자치도"],
    "전남": ["전남", "전라남도"],
    "경북": ["경북", "경상북도"],
    "경남": ["경남", "경상남도", "창원"],
    "제주": ["제주"],
}


def raw_file(dataset_id: str, suffix: str | None = None) -> Path:
    files = sorted(RAW_DIR.glob(f"{dataset_id}_*"))
    if suffix:
        files = [p for p in files if p.suffix.lower() == suffix]
    if not files:
        raise FileNotFoundError(dataset_id)
    return files[0]


def read_csv_any(path: Path) -> list[dict[str, str]]:
    for enc in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                return list(csv.DictReader(f))
        except UnicodeDecodeError:
            continue
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        return list(csv.DictReader(f))


def to_number(value: str | int | float | None) -> float:
    if value is None:
        return 0.0
    text = str(value)
    text = re.sub(r"[^0-9.\-]", "", text)
    return float(text) if text else 0.0


def region_from_text(text: str) -> str | None:
    for region, aliases in REGION_ALIASES.items():
        if any(alias in text for alias in aliases):
            return region
    return None


def normalize(value: float, min_v: float, max_v: float, low: float, high: float) -> float:
    if math.isclose(max_v, min_v):
        return round((low + high) / 2, 3)
    return round(low + (value - min_v) / (max_v - min_v) * (high - low), 3)


def build_region_counts() -> dict[str, Counter]:
    counts = {
        "centers": Counter(),
        "apprenticeshipCompanies": Counter(),
        "subsidyOperators": Counter(),
        "localProjects": Counter(),
    }

    for row in read_csv_any(raw_file("15066368", ".csv")):
        region = region_from_text(" ".join(row.values()))
        if region:
            counts["centers"][region] += 1

    for row in read_csv_any(raw_file("15118732", ".csv")):
        region = region_from_text(" ".join(row.values()))
        if region:
            counts["apprenticeshipCompanies"][region] += 1

    for row in read_csv_any(raw_file("15119494", ".csv")):
        region = region_from_text(" ".join(row.values()))
        if region:
            counts["subsidyOperators"][region] += 1

    for row in read_csv_any(raw_file("15120704", ".csv")):
        region = region_from_text(" ".join(row.values()))
        if region:
            counts["localProjects"][region] += 1

    return counts


def latest_minimum_wage() -> dict:
    rows = read_csv_any(raw_file("15068774", ".csv"))
    rows = sorted(rows, key=lambda r: to_number(r.get("연도")), reverse=True)
    latest = rows[0]
    return {
        "year": int(to_number(latest.get("연도"))),
        "hourlyWage": int(to_number(latest.get("시간급"))),
        "rowCount": len(rows),
    }


def subsidy_performance() -> dict:
    rows = read_csv_any(raw_file("15154122", ".csv"))
    parsed = []
    for row in rows:
        parsed.append({
            "year": int(to_number(row.get("연도"))),
            "companies": int(to_number(row.get("기업(개)"))),
            "youth": int(to_number(row.get("청년(명)"))),
        })
    parsed = sorted(parsed, key=lambda r: r["year"])
    latest = parsed[-1]
    max_youth = max(r["youth"] for r in parsed) or 1
    return {
        "latest": latest,
        "history": parsed,
        "subsidyEffect": round(0.045 + 0.035 * latest["youth"] / max_youth, 3),
    }


def training_it_status() -> dict:
    rows = read_csv_any(raw_file("15139460", ".csv"))
    parsed = []
    for row in rows:
        parsed.append({
            "year": int(to_number(row.get("훈련시작년"))),
            "participants": int(to_number(row.get("실시인원(명)"))),
        })
    parsed = sorted(parsed, key=lambda r: r["year"])
    latest = parsed[-1]
    max_participants = max(r["participants"] for r in parsed) or 1
    return {
        "latest": latest,
        "history": parsed,
        "trainingBaseEffect": round(0.055 + 0.045 * latest["participants"] / max_participants, 3),
    }


def unemployment_pressure() -> dict:
    rows = read_csv_any(raw_file("15029716", ".csv"))
    parsed = []
    for row in rows:
        if row.get("유형") == "구직급여":
            parsed.append({
                "year": int(to_number(row.get("연도"))),
                "recipients": int(to_number(row.get("지급자수"))),
                "amountMillionWon": int(to_number(row.get("지급액(백만원)"))),
            })
    parsed = sorted(parsed, key=lambda r: r["year"])
    latest = parsed[-1]
    max_recipients = max(r["recipients"] for r in parsed) or 1
    return {
        "latest": latest,
        "history": parsed,
        "jobSearchPressure": round(0.04 + 0.05 * latest["recipients"] / max_recipients, 3),
    }


def benefit_days() -> dict:
    rows = read_csv_any(raw_file("15068742", ".csv"))
    values = []
    for row in rows:
        for key, value in row.items():
            if "피보험기간" in key:
                n = to_number(value)
                if n:
                    values.append(n)
    return {
        "minDays": int(min(values)),
        "maxDays": int(max(values)),
        "averageDays": round(sum(values) / len(values), 1),
        "rowCount": len(rows),
    }


def wage_by_job(existing_jobs: dict) -> tuple[dict, dict]:
    rows = read_csv_any(raw_file("15133009", ".csv"))
    by_industry = {row["산업대분류"]: to_number(row.get("연간급여중위값") or row.get("연간급여평균값")) for row in rows}
    median_all = sorted(by_industry.values())[len(by_industry) // 2]
    mapping = {
        "데이터 분석": "통신업",
        "개발": "통신업",
        "사무": "사업서비스업",
        "마케팅": "도매·소매업",
        "디자인": "오락·문화·운동관련 서비스업",
    }
    updated = {}
    for job, profile in existing_jobs.items():
        industry = mapping.get(job, "사업서비스업")
        median = by_industry.get(industry, median_all)
        base = max(210, min(360, round((median / median_all) * 255)))
        updated[job] = {**profile, "baseWage": base, "sourceIndustry": industry}
    return updated, by_industry


def ncs_summary() -> dict:
    path = raw_file("15106008", ".xlsx")
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        non_empty_rows = 0
        sample = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(v).strip() for v in row if v not in (None, "")]
            if values:
                non_empty_rows += 1
                if len(sample) < 5:
                    sample.append(values[:6])
        return {"format": "xlsx", "sheet": sheet.title, "nonEmptyRows": non_empty_rows, "sample": sample}
    except Exception:
        with zipfile.ZipFile(path) as zf:
            names = zf.namelist()
            text_files = [name for name in names if name.lower().endswith((".xml", ".txt"))]
        return {
            "format": "hwp_zip",
            "fileCount": len(names),
            "textFileCount": len(text_files),
            "sampleFiles": text_files[:8],
        }


def build_market(existing_market: dict, counts: dict[str, Counter]) -> dict:
    support_score = {}
    for region in REGIONS:
        support_score[region] = (
            counts["centers"][region] * 2.0
            + counts["subsidyOperators"][region] * 1.6
            + counts["apprenticeshipCompanies"][region] * 0.045
            + counts["localProjects"][region] * 0.55
        )

    min_score = min(support_score.values())
    max_score = max(support_score.values())
    market = {}
    for region in REGIONS:
        base = existing_market.get(region, {"demand": 0.5, "competition": 0.5, "wage": 0.95})
        support = normalize(support_score[region], min_score, max_score, 0.35, 0.82)
        demand = round(base["demand"] * 0.55 + support * 0.45, 3)
        competition = round(base["competition"] * 0.7 + normalize(counts["subsidyOperators"][region], 0, max(counts["subsidyOperators"].values() or [1]), 0.35, 0.72) * 0.3, 3)
        access = normalize(counts["centers"][region], 0, max(counts["centers"].values() or [1]), 0.3, 0.9)
        market[region] = {
            "demand": max(0.25, min(0.9, demand)),
            "competition": max(0.25, min(0.85, competition)),
            "wage": base["wage"],
            "supportAccess": access,
            "rawCounts": {
                "centers": counts["centers"][region],
                "apprenticeshipCompanies": counts["apprenticeshipCompanies"][region],
                "subsidyOperators": counts["subsidyOperators"][region],
                "localProjects": counts["localProjects"][region],
            },
        }
    return market


def main() -> None:
    data = json.loads(DATASETS_JSON.read_text(encoding="utf-8"))
    counts = build_region_counts()
    wage_jobs, wage_industries = wage_by_job(data["jobProfiles"])
    min_wage = latest_minimum_wage()
    subsidy = subsidy_performance()
    training = training_it_status()
    pressure = unemployment_pressure()
    days = benefit_days()
    ncs = ncs_summary()
    market = build_market(data["market"], counts)

    policy_effects = {
        **data["policyEffects"],
        "counseling": round(0.04 + 0.04 * (sum(counts["centers"].values()) / max(1, sum(counts["centers"].values()) + 60)), 3),
        "trainingBase": training["trainingBaseEffect"],
        "subsidy": subsidy["subsidyEffect"],
        "integratedSynergy": 0.05,
        "distancePenalty": round(0.04 - min(m["supportAccess"] for m in market.values()) * 0.02, 3),
        "jobSearchPressure": pressure["jobSearchPressure"],
    }

    summary = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "method": "20개 고용노동부 원자료를 내려받아 CSV, XLSX, HTML형 원문을 적재하고 MVP 계산용 변수로 요약",
        "rawManifest": "public/data/raw/manifest.json",
        "minimumWage": min_wage,
        "subsidyPerformance": subsidy,
        "trainingItStatus": training,
        "unemploymentPressure": pressure,
        "benefitDays": days,
        "ncs": ncs,
        "wageByIndustryMedian": wage_industries,
        "regionCounts": {name: dict(counter) for name, counter in counts.items()},
        "market": market,
        "jobProfiles": wage_jobs,
        "policyEffects": policy_effects,
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    data["source"] = {
        **data["source"],
        "rawDataLoaded": True,
        "rawDatasetCount": 20,
        "rawManifest": "public/data/raw/manifest.json",
        "analysisSummary": "public/data/moel_analysis_summary.json",
        "analysisGeneratedAt": summary["generatedAt"],
    }
    data["market"] = market
    data["jobProfiles"] = wage_jobs
    data["policyEffects"] = policy_effects
    DATASETS_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"summary: {SUMMARY_JSON}")
    print("minimum wage", min_wage)
    print("policy effects", policy_effects)
    print("top support regions", sorted(((r, market[r]['rawCounts']) for r in REGIONS), key=lambda x: sum(x[1].values()), reverse=True)[:5])


if __name__ == "__main__":
    main()
