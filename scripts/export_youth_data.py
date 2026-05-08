from pathlib import Path
import json
from openpyxl import load_workbook

BASE = Path(__file__).resolve().parents[1]
XLSX = next((BASE / "outputs" / "moel-public-data").glob("*.xlsx"))
OUT = BASE / "public" / "data" / "datasets.json"

required = {
    "15068741": ("고용24구인구직 취업동향", "지역·직무별 수요와 바로 구직 경로의 취업확률 기준선"),
    "15033150": ("한권으로 통하는 대한민국 청년지원 프로그램 정보", "청년에게 가능한 정책 경로 후보 생성"),
    "3038234": ("국민취업지원제도 운영기관", "상담기관 연결과 상담 경로 효과 산정"),
    "15066368": ("고용복지플러스센터 현황", "오프라인 전달체계와 기관 접근성 계산"),
    "15029716": ("구직급여 수급 현황", "실업 상태 기준선과 구직기간 제약 추정"),
    "15068742": ("구직급여 소정급여일수", "구직급여 가능 기간에 따른 압박도 계산"),
    "15154122": ("청년일자리도약장려금 사업추진실적", "장려금 기업 경로의 채용확률 보정"),
    "15119494": ("청년일자리도약장려금 사업 운영기관", "장려금 가능 기업·운영기관 연결"),
    "15113450": ("국민내일배움카드 발급 현황", "훈련 참여 가능성 보정"),
    "15139460": ("국민내일배움카드 정보통신분야 실시현황", "IT·데이터 직무 훈련 경로 보정"),
    "15117780": ("직업능력개발 사업현황", "훈련사업 유형과 정책수단 확장"),
    "15106008": ("국가직무능력표준(NCS) 개발 개선 현황", "희망직무와 보유역량 간 스킬갭 계산"),
    "15118639": ("국가기술자격 통계연보", "자격취득 경로의 효용 추정"),
    "3069913": ("직종별사업체노동력조사", "직종별 빈일자리·부족인원 기준선"),
    "3069922": ("지역별사업체노동력조사", "지역 노동수요와 경기 변화 반영"),
    "3069910": ("사업체노동력조사", "산업별 고용수요 보정"),
    "15133009": ("300명이상 사업체의 산업대분류별 임금수준", "예상 초임·임금 경로 보정"),
    "15068774": ("연도별 최저임금", "예상임금 하한선 설정"),
    "15118732": ("일학습병행 참여기업", "훈련+취업 결합 경로 후보"),
    "15120704": ("지역산업맞춤형 일자리지원사업 접수처", "지역 맞춤사업 경로 확장"),
}

sample_market = {
    "서울": {"demand": 0.72, "competition": 0.62, "wage": 1.12},
    "경기": {"demand": 0.68, "competition": 0.58, "wage": 1.05},
    "부산": {"demand": 0.53, "competition": 0.55, "wage": 0.94},
    "대구": {"demand": 0.49, "competition": 0.51, "wage": 0.91},
    "광주": {"demand": 0.46, "competition": 0.49, "wage": 0.90},
    "대전": {"demand": 0.56, "competition": 0.50, "wage": 0.97},
    "인천": {"demand": 0.57, "competition": 0.53, "wage": 0.98},
    "울산": {"demand": 0.50, "competition": 0.44, "wage": 1.02},
    "세종": {"demand": 0.52, "competition": 0.47, "wage": 1.00},
    "강원": {"demand": 0.39, "competition": 0.38, "wage": 0.86},
    "충북": {"demand": 0.48, "competition": 0.42, "wage": 0.91},
    "충남": {"demand": 0.51, "competition": 0.43, "wage": 0.95},
    "전북": {"demand": 0.42, "competition": 0.39, "wage": 0.86},
    "전남": {"demand": 0.40, "competition": 0.37, "wage": 0.85},
    "경북": {"demand": 0.45, "competition": 0.39, "wage": 0.88},
    "경남": {"demand": 0.47, "competition": 0.41, "wage": 0.90},
    "제주": {"demand": 0.41, "competition": 0.44, "wage": 0.87},
}

job_profiles = {
    "데이터 분석": {
        "ncs": "정보기술 > 빅데이터분석",
        "required": ["Excel", "SQL", "Python", "데이터 시각화", "기초 통계", "문제정의"],
        "adjacent": ["데이터 운영", "BI 리포팅", "사무자동화", "마케팅 데이터 보조"],
        "baseWage": 255,
        "trainingBoost": 0.13,
    },
    "사무": {
        "ncs": "경영·회계·사무 > 총무·인사",
        "required": ["Excel", "문서작성", "커뮤니케이션", "일정관리", "자료정리"],
        "adjacent": ["영업지원", "HR 운영", "총무", "고객운영"],
        "baseWage": 225,
        "trainingBoost": 0.07,
    },
    "마케팅": {
        "ncs": "영업판매 > 마케팅",
        "required": ["콘텐츠 기획", "Excel", "데이터 리터러시", "SNS 운영", "문서작성"],
        "adjacent": ["콘텐츠 운영", "CRM 보조", "광고 운영", "영업지원"],
        "baseWage": 235,
        "trainingBoost": 0.09,
    },
    "개발": {
        "ncs": "정보기술 > 응용SW엔지니어링",
        "required": ["JavaScript", "Python", "Git", "SQL", "API 이해", "문제해결"],
        "adjacent": ["QA", "웹 퍼블리싱", "데이터 엔지니어링 보조", "IT 운영"],
        "baseWage": 275,
        "trainingBoost": 0.15,
    },
    "디자인": {
        "ncs": "문화·예술·디자인 > 디자인",
        "required": ["Figma", "포트폴리오", "사용자 이해", "시각디자인", "커뮤니케이션"],
        "adjacent": ["콘텐츠 디자인", "서비스 운영", "UX 리서치 보조", "마케팅 디자인"],
        "baseWage": 230,
        "trainingBoost": 0.08,
    },
}

wb = load_workbook(XLSX, read_only=True, data_only=True)
ws = wb["목록"]
rows = list(ws.iter_rows(values_only=True))[1:]
catalog = {}
for r in rows:
    data_id = str(r[1])
    if data_id in required:
        catalog[data_id] = {
            "no": r[0],
            "id": data_id,
            "title": r[2],
            "description": r[3],
            "category": r[4],
            "keywords": r[11] or "",
            "role": required[data_id][1],
        }

payload = {
    "source": {
        "catalogFile": XLSX.name,
        "totalCatalogRows": len(rows),
        "selectedRows": len(catalog),
        "generatedAt": "2026-05-08",
    },
    "datasets": list(catalog.values()),
    "market": sample_market,
    "jobProfiles": job_profiles,
    "policyEffects": {
        "counseling": 0.06,
        "trainingBase": 0.08,
        "subsidy": 0.07,
        "integratedSynergy": 0.05,
        "distancePenalty": 0.03,
    },
}

OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(OUT)
