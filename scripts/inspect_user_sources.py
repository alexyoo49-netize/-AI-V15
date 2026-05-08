from pathlib import Path
from openpyxl import load_workbook
import csv
import sys

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path.home() / "Desktop" / "데이터 셋 모음 고용노동"

NEEDLES = [
    "내일배움",
    "실업자훈련",
    "훈련기관",
    "국민취업지원제도 운영기관",
    "NCS",
    "청년일자리도약장려금",
    "구인구직취업현황",
]

for p in ROOT.rglob("*"):
    if not p.is_file():
        continue
    if not any(n in p.name for n in NEEDLES):
        continue
    print("\nFILE", p.name, "PARENT", p.parent.name, "SIZE", p.stat().st_size)
    if p.suffix.lower() == ".csv":
        for enc in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                with p.open(encoding=enc, newline="") as f:
                    reader = csv.reader(f)
                    for i, row in zip(range(1, 8), reader):
                        print(i, row[:12])
                print("ENC", enc)
                break
            except Exception:
                pass
        continue
    if p.suffix.lower() != ".xlsx":
        continue
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        print("SHEETS", wb.sheetnames)
        for ws in wb.worksheets[:2]:
            print("SHEET", ws.title, ws.max_row, ws.max_column)
            for i, row in zip(range(1, 16), ws.iter_rows(values_only=True)):
                print(i, list(row[:12]))
    except Exception as e:
        print("ERR", repr(e))
