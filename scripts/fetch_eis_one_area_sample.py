from __future__ import annotations

import json
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parents[1] / "outputs" / "eis-api-samples"
AREA_CODE = "11110"
AREA_NAME = "서울특별시 종로구"
MONTH = "202401"
SEX_CODE = "M"
SEX_NAME = "남자"
AGE_CODE = "03"
AGE_NAME = "25~29세"
DISPLAY = 9999


def fetch_xml(page: int, display: int) -> bytes:
    params = {
        "apiSecd": "OPIA",
        "rsdAreaCd": AREA_CODE,
        "sxdsCd": SEX_CODE,
        "ageCd": AGE_CODE,
        "rernSecd": "XML",
        "closStdrYm": MONTH,
        "bgnPage": str(page),
        "display": str(display),
    }
    url = "https://eis.work24.go.kr/opi/joApi.do?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=40) as res:
        return res.read()


def parse_response(content: bytes) -> tuple[int, list[dict[str, str]]]:
    text = content.decode("euc-kr", "replace")
    text = text.replace('encoding="EUC-KR"', 'encoding="UTF-8"')
    root = ET.fromstring(text.encode("utf-8"))
    cnt = int(root.findtext("rqst-cnt") or 0)
    records = []
    for node in root.findall("./rqst-list/rqst"):
        row = {}
        for child in list(node):
            row[child.tag] = child.text or ""
        records.append(row)
    return cnt, records


def autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 34)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xml = fetch_xml(page=1, display=DISPLAY)
    total_count, records = parse_response(xml)

    out_xml = OUT_DIR / f"eis_joApi_{AREA_CODE}_{MONTH}_{SEX_CODE}_{AGE_CODE}.xml"
    out_json = OUT_DIR / f"eis_joApi_{AREA_CODE}_{MONTH}_{SEX_CODE}_{AGE_CODE}.json"
    out_xlsx = OUT_DIR / f"eis_joApi_{AREA_CODE}_{MONTH}_{SEX_CODE}_{AGE_CODE}.xlsx"
    out_xml.write_bytes(xml)
    out_json.write_text(json.dumps({
        "fetchedAt": datetime.now().isoformat(timespec="seconds"),
        "areaCode": AREA_CODE,
        "areaName": AREA_NAME,
        "month": MONTH,
        "sexCode": SEX_CODE,
        "sexName": SEX_NAME,
        "ageCode": AGE_CODE,
        "ageName": AGE_NAME,
        "page": 1,
        "display": DISPLAY,
        "totalCount": total_count,
        "returnedCount": len(records),
        "records": records,
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "EIS API 샘플"
    ws.append(["항목", "값"])
    meta = [
        ["API", "https://eis.work24.go.kr/opi/joApi.do"],
        ["지역코드", AREA_CODE],
        ["지역명", AREA_NAME],
        ["기준월", MONTH],
        ["성별", f"{SEX_CODE} {SEX_NAME}"],
        ["연령", f"{AGE_CODE} {AGE_NAME}"],
        ["페이지번호 bgnPage", 1],
        ["출력건수 display", DISPLAY],
        ["API 총 건수 rqst-cnt", total_count],
        ["이번 파일 수록 건수", len(records)],
        ["설명", "display는 한 페이지에 받을 건수다. rqst-cnt가 display보다 작으면 bgnPage 1만 호출하면 된다. rqst-cnt가 display보다 크면 다음 페이지를 추가 호출한다."],
    ]
    for row in meta:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003675")
    autosize(ws)

    data_ws = wb.create_sheet("records")
    keys = sorted({key for row in records for key in row.keys()})
    preferred = [
        "dwClosYm", "ctpvCd", "ctpvCdNm", "rsdAreaCd", "rsdAreaCdNm",
        "sxdsCd", "sxdsCdNm", "ageCd", "ageCdNm",
        "jsfcVerCdNm", "jsfcLrclCdNm", "jsfcMdclCdNm",
        "newJoNmpr", "newJhntNmpr", "empmCt",
    ]
    headers = [key for key in preferred if key in keys] + [key for key in keys if key not in preferred]
    data_ws.append(headers)
    for row in records:
        data_ws.append([row.get(key, "") for key in headers])
    for cell in data_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003675")
        cell.alignment = Alignment(horizontal="center")
    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = data_ws.dimensions
    autosize(data_ws)

    wb.save(out_xlsx)
    print(out_xlsx)
    print(f"total_count={total_count} returned={len(records)} display={DISPLAY}")


if __name__ == "__main__":
    main()
