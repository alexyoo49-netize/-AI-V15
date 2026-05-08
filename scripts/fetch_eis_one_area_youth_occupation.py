from __future__ import annotations

import json
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


OUT_DIR = Path(__file__).resolve().parents[1] / "outputs" / "eis-api-samples"
AREA_CODE = "11110"
AREA_NAME = "서울특별시 종로구"
MONTHS = [f"{year}{month:02d}" for year in range(2021, 2027) for month in range(1, 13) if "202104" <= f"{year}{month:02d}" <= "202603"]
SEXES = [("M", "남자"), ("F", "여자")]
AGES = [("02", "20~24세"), ("03", "25~29세"), ("04", "30~34세")]
DISPLAY = 9999


def fetch_xml(month: str, sex_code: str, age_code: str, page: int = 1) -> bytes:
    params = {
        "apiSecd": "OPIA",
        "rsdAreaCd": AREA_CODE,
        "sxdsCd": sex_code,
        "ageCd": age_code,
        "rernSecd": "XML",
        "closStdrYm": month,
        "bgnPage": str(page),
        "display": str(DISPLAY),
    }
    url = "https://eis.work24.go.kr/opi/joApi.do?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=40) as res:
        return res.read()


def parse_response(content: bytes) -> tuple[int, list[dict[str, str]]]:
    text = content.decode("euc-kr", "replace").replace('encoding="EUC-KR"', 'encoding="UTF-8"')
    if "</rqstApi>" in text:
        text = text[: text.index("</rqstApi>") + len("</rqstApi>")]
    elif "</baroone>" in text:
        return 0, []
    root = ET.fromstring(text.encode("utf-8"))
    cnt = int(root.findtext("rqst-cnt") or 0)
    rows = []
    for node in root.findall("./rqst-list/rqst"):
        rows.append({child.tag: child.text or "" for child in list(node)})
    return cnt, rows


def num(value: str) -> int:
    try:
        return int(float(value or 0))
    except ValueError:
        return 0


def autosize(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = max(len(str(cell.value or "")) for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003675")
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    raw_rows = []
    calls = []

    for month in MONTHS:
        for sex_code, sex_name in SEXES:
            for age_code, age_name in AGES:
                xml = fetch_xml(month, sex_code, age_code, 1)
                total, rows = parse_response(xml)
                calls.append({
                    "month": month,
                    "sexCode": sex_code,
                    "sexName": sex_name,
                    "ageCode": age_code,
                    "ageName": age_name,
                    "totalCount": total,
                    "returnedCount": len(rows),
                    "display": DISPLAY,
                })
                for row in rows:
                    raw_rows.append({
                        "기준월": row.get("dwClosYm", month),
                        "연도": row.get("dwClosYm", month)[:4],
                        "시도": row.get("ctpvCdNm", ""),
                        "지역코드": row.get("rsdAreaCd", AREA_CODE),
                        "지역명": row.get("rsdAreaCdNm", AREA_NAME),
                        "성별": row.get("sxdsCdNm", sex_name),
                        "연령": row.get("ageCdNm", age_name),
                        "직종코드": row.get("wnetJsfcLrclCd", ""),
                        "직종명": row.get("wnetJsfcLrclCdNm", ""),
                        "신규구인인원": num(row.get("newJoNmpr")),
                        "신규구직건수": num(row.get("newJhntNmpr")),
                        "취업건수": num(row.get("empmCt")),
                        "유효구인인원": num(row.get("valdJoNmpr")),
                    })
                time.sleep(0.08)

    monthly = defaultdict(lambda: {"신규구인인원": 0, "신규구직건수": 0, "취업건수": 0, "유효구인인원": 0})
    yearly = defaultdict(lambda: {"신규구인인원": 0, "신규구직건수": 0, "취업건수": 0, "유효구인인원": 0})
    for row in raw_rows:
        m_key = (row["기준월"], row["시도"], row["지역명"], row["직종코드"], row["직종명"])
        y_key = (row["연도"], row["시도"], row["지역명"], row["직종코드"], row["직종명"])
        for field in ["신규구인인원", "신규구직건수", "취업건수", "유효구인인원"]:
            monthly[m_key][field] += row[field]
            yearly[y_key][field] += row[field]

    out_xlsx = OUT_DIR / f"eis_종로구_청년_직종별_구인구직취업_202104_202603.xlsx"
    out_json = OUT_DIR / f"eis_종로구_청년_직종별_구인구직취업_202104_202603.json"
    out_json.write_text(json.dumps({
        "fetchedAt": datetime.now().isoformat(timespec="seconds"),
        "areaCode": AREA_CODE,
        "areaName": AREA_NAME,
        "months": [MONTHS[0], MONTHS[-1]],
        "sexes": SEXES,
        "ages": AGES,
        "display": DISPLAY,
        "apiCalls": calls,
        "rawRecordCount": len(raw_rows),
    }, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.title = "설명"
    ws.append(["항목", "값"])
    for row in [
        ["API", "https://eis.work24.go.kr/opi/joApi.do"],
        ["지역", f"{AREA_CODE} {AREA_NAME}"],
        ["기간", f"{MONTHS[0]}부터 {MONTHS[-1]}까지"],
        ["대상", "20~34세 남자와 여자 합산"],
        ["출력건수 display", DISPLAY],
        ["API 호출 수", len(calls)],
        ["원시 행 수", len(raw_rows)],
        ["주의", "이 파일은 시군구 하나의 검증 샘플이다. 전국 구축 시 같은 방식으로 시군구 코드를 반복 호출하고 시도 단위로 합산한다."],
    ]:
        ws.append(row)
    style_header(ws)
    autosize(ws)

    monthly_ws = wb.create_sheet("월별_직종별_집계")
    monthly_ws.append(["기준월", "시도", "지역명", "직종코드", "직종명", "신규구인인원", "신규구직건수", "취업건수", "유효구인인원", "취업률_취업건수_구직건수"])
    for key, vals in sorted(monthly.items()):
        구직 = vals["신규구직건수"]
        취업 = vals["취업건수"]
        monthly_ws.append([*key, vals["신규구인인원"], 구직, 취업, vals["유효구인인원"], round(취업 / 구직, 4) if 구직 else 0])
    style_header(monthly_ws)
    autosize(monthly_ws)

    yearly_ws = wb.create_sheet("연도별_직종별_집계")
    yearly_ws.append(["연도", "시도", "지역명", "직종코드", "직종명", "신규구인인원", "신규구직건수", "취업건수", "유효구인인원", "취업률_취업건수_구직건수"])
    for key, vals in sorted(yearly.items()):
        구직 = vals["신규구직건수"]
        취업 = vals["취업건수"]
        yearly_ws.append([*key, vals["신규구인인원"], 구직, 취업, vals["유효구인인원"], round(취업 / 구직, 4) if 구직 else 0])
    style_header(yearly_ws)
    autosize(yearly_ws)

    raw_ws = wb.create_sheet("API_원시행")
    headers = ["기준월", "연도", "시도", "지역코드", "지역명", "성별", "연령", "직종코드", "직종명", "신규구인인원", "신규구직건수", "취업건수", "유효구인인원"]
    raw_ws.append(headers)
    for row in raw_rows:
        raw_ws.append([row[h] for h in headers])
    style_header(raw_ws)
    autosize(raw_ws)

    calls_ws = wb.create_sheet("API_호출로그")
    calls_ws.append(["기준월", "성별코드", "성별", "연령코드", "연령", "전체건수", "반환건수", "display"])
    for call in calls:
        calls_ws.append([call["month"], call["sexCode"], call["sexName"], call["ageCode"], call["ageName"], call["totalCount"], call["returnedCount"], call["display"]])
    style_header(calls_ws)
    autosize(calls_ws)

    wb.save(out_xlsx)
    print(out_xlsx)
    print(f"calls={len(calls)} raw_rows={len(raw_rows)} monthly_rows={len(monthly)} yearly_rows={len(yearly)}")


if __name__ == "__main__":
    main()
